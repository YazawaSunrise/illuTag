from __future__ import annotations

import argparse
import base64
import csv
import json
import os
import re
import shutil
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook


OUTPUT_DIR_NAME = "\u5168\u91cf\u7ffb\u8bd1"
HAS_HAN = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]")
HAS_KANA = re.compile(r"[\u3040-\u30ff\uff66-\uff9f]")
HAS_HANGUL = re.compile(r"[\uac00-\ud7af]")

# Small heuristics only decide which API candidate is safer to put in column 1.
# All ambiguous Han-only candidates are still preserved in zh_name_2 for review.
JP_ONLY_HINTS = set("兎辻込働峠畑姫凧匂躾栃榊雫咲")
ZH_HINTS = set(
    "兔条纹饰装换衫裙裤袜发脸胸腿脚眼嘴唇牙耳尾角翼毛皮肤"
    "黑白红蓝绿黄紫粉橙棕灰金银色长短大小高低上下左右前后"
    "官方服衣帽带领结蝴蝶结泳装制服水手校服围巾手套靴鞋"
    "透明湿裸全半单双多女性男性少女少年萝莉正太动物猫狗狐"
)
SIMPLIFIED_HINTS = set(
    "个为义乌乐习乡书买争于亚产亲亿仅从仓们优会伞传伤体余"
    "侧侦俩俦债倾偿儿克兑兰关兴养兽内军农冲决况净准几击划"
    "刘则刚创删别制刹剂剑剧劝办务动励劲劳势勋区华协单卢卫"
    "厅历厉压县参双发变叠叶号叹后向吨启员响团园围国图圆圣"
    "场块坚坛坝垫处备复头夹奖妆妇妈孙学宝实宠审宪宫宽宾对"
    "导将尔尘尝尽层届属屡岁岂岗岛岭岳币帅师帐帘带帮广庄庆"
    "库应庙废开异弃张弥弯弹强归当录彻径忆忧态怀总恋恶恼悦"
    "悬惊惨惩惯愿懒戏战户扎执扩扫扬扰抚抛抢护报担拟拢拥拦"
    "拨择挂挚挛挥挤换据捡掀控掺揽搁搂搅摄摆摇撤播撵敌敛数"
    "断无旧时旷显晓暂术机杀杂权条来极构枢枪枫柜标栈栋栏树"
    "样档桥检楼概横樱橱欢欧歼残殴毁毕毙气汉汤沟没沧洁洒测"
    "济浏浓涂涛涝渐渔温游湾湿滚满滥滨滩灭灯灵灾炉点炼热爱"
    "爷牵状独狭狮猎猪献玛环现电画畅疗疯痒监盖盘矫矿码砖硕"
    "确礼离种称积稳窍窜竞笔筛简类纠红纤约级纪纱纲纳纵纷纸"
    "纹纺纽线练组细终绍经绑结绕绘给络绝统绣继续绳维绵绿编"
    "缘缩网罚罢职联聪肤胜胶胸脸脱艺节苏范荣药莱获营蓝蔑虏"
    "虑虚虽虾蛮 补装裤见观规视觉览触计订认讨让训议记讲许论"
    "设访证评识诉词试诗诚话询该详语误诱说请诸诺读调谈谋谢"
    "贝负贡财责贤败货质贩贪贫贬购贵费贺贼资赋赌赏赛赞赠赢"
    "赵趋跃践车轨转轮软轰轴轻载较辅辆辉辑输达迁过运还进远"
    "连迟选递遗遥邮邻郁郑配里鉴钢钥钱钻铁铃铜铝银链铺销锁"
    "锅错键镇镜长门问闲间闷闻队阳阴阵阶际陆陈限险难静页顶"
    "项顺须顾预领频题颜额风飞饭饮饰饼馆马驱验骑骗骚鱼鸟鸡"
    "鹅麦黄黑齐齿"
)


def read_csv_dict(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv_dict(path: Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def load_dictionary(path: Path) -> dict[str, str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["general"] if "general" in wb.sheetnames else wb[wb.sheetnames[0]]
    result: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tag = row[2] if len(row) > 2 else None
        zh = row[3] if len(row) > 3 else None
        if tag is None or zh is None:
            continue
        tag_s = str(tag).strip()
        zh_s = str(zh).strip()
        if tag_s and zh_s:
            result[tag_s] = zh_s
    return result


def flatten_other_names(value: Any) -> list[str]:
    if isinstance(value, list):
        raw = value
    elif isinstance(value, str):
        raw = re.split(r"[,;\n\r]+", value)
    else:
        raw = []
    names: list[str] = []
    seen = set()
    for item in raw:
        name = str(item).strip()
        if name and name not in seen:
            seen.add(name)
            names.append(name)
    return names


def is_han_only_candidate(text: str) -> bool:
    return bool(HAS_HAN.search(text)) and not HAS_KANA.search(text) and not HAS_HANGUL.search(text)


def candidate_score(text: str) -> tuple[int, int, int]:
    simplified = sum(1 for ch in text if ch in SIMPLIFIED_HINTS)
    zh_hint = sum(1 for ch in text if ch in ZH_HINTS)
    jp_penalty = sum(1 for ch in text if ch in JP_ONLY_HINTS)
    han_count = len(HAS_HAN.findall(text))
    ascii_penalty = len(re.findall(r"[A-Za-z]", text))
    length_penalty = abs(len(text) - 4)
    return (simplified * 3 + zh_hint * 2 + han_count - jp_penalty * 4 - ascii_penalty, -length_penalty, -len(text))


def choose_api_translation(other_names: Any) -> tuple[str, str, str]:
    candidates = [name for name in flatten_other_names(other_names) if is_han_only_candidate(name)]
    if not candidates:
        return "", "", "no_han_candidate"
    candidates.sort(key=candidate_score, reverse=True)
    best = candidates[0]
    mixed_or_ambiguous = len(candidates) > 1 or any(ch in JP_ONLY_HINTS for ch in best)
    zh_name_2 = "; ".join(candidates[:12]) if mixed_or_ambiguous else ""
    note = "api_han_candidates" if mixed_or_ambiguous else "api_single_han_candidate"
    return best, zh_name_2, note


def request_json(
    session: requests.Session,
    url: str,
    login: str,
    api_key: str,
    timeout: int,
) -> tuple[int, Any, str]:
    headers = {
        "Accept": "application/json",
        "User-Agent": "illuTag-full-translator/0.1 (official Danbooru API lookup)",
    }
    auth = (login, api_key) if login and api_key else None
    try:
        response = session.get(url, headers=headers, auth=auth, timeout=timeout)
        if response.status_code == 200:
            return response.status_code, response.json(), ""
        return response.status_code, None, response.text[:500]
    except Exception as exc:
        return 0, None, str(exc)


def wiki_urls(base_url: str, tag: str) -> list[str]:
    base = base_url.rstrip("/")
    quoted = urllib.parse.quote(tag, safe="")
    query = urllib.parse.urlencode({"search[title]": tag})
    return [f"{base}/wiki_pages/{quoted}.json", f"{base}/wiki_pages.json?{query}"]


def load_cache(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def save_cache(path: Path, cache: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2, sort_keys=True)
    for attempt in range(5):
        try:
            tmp.replace(path)
            return
        except PermissionError:
            if attempt == 4:
                raise
            time.sleep(0.5)


def fetch_wiki(
    session: requests.Session,
    tag: str,
    args: argparse.Namespace,
    cache: dict[str, Any],
) -> dict[str, Any]:
    if tag in cache and not args.refresh:
        return cache[tag]

    last_error = ""
    for url in wiki_urls(args.base_url, tag):
        status, payload, error = request_json(session, url, args.login, args.api_key, args.timeout)
        auth_fallback = False
        if status == 403 and args.login and args.api_key and "authenticate_api_key" in error:
            # Public wiki reads do not require auth. If the provided key is bad,
            # fall back to anonymous official API access instead of poisoning the run.
            auth_fallback = True
            status, payload, error = request_json(session, url, "", "", args.timeout)
        if status == 200:
            if isinstance(payload, list):
                payload = payload[0] if payload else None
            record = {
                "status": "ok" if payload else "not_found",
                "http_status": status,
                "url": url,
                "payload": payload,
                "error": "auth_failed_used_anonymous" if auth_fallback else "",
            }
            cache[tag] = record
            return record
        if status == 404:
            last_error = "404 not found"
            continue
        if status == 429:
            time.sleep(args.retry_after)
            continue
        last_error = f"HTTP {status}: {error}"
        break

    record = {
        "status": "error" if last_error else "not_found",
        "http_status": 0,
        "url": "",
        "payload": None,
        "error": last_error,
    }
    cache[tag] = record
    return record


def build_row(row: dict[str, str], dictionary: dict[str, str], wiki: dict[str, Any] | None) -> dict[str, str]:
    tag = row["name"]
    base = {
        "tag_id": row["tag_id"],
        "name": tag,
        "category": row["category"],
        "count": row["count"],
        "zh_name": "",
        "zh_name_2": "",
        "source": "",
        "confidence": "",
        "note": "",
        "donmai_url": f"https://danbooru.donmai.us/wiki_pages/{urllib.parse.quote(tag, safe='')}",
        "api_status": "",
    }
    if tag in dictionary:
        base.update({"zh_name": dictionary[tag], "source": "dictionary01", "confidence": "high"})
        return base

    if not wiki:
        base.update({"source": "not_requested", "confidence": "low", "note": "api_lookup_skipped"})
        return base

    payload = wiki.get("payload") or {}
    zh_name, zh_name_2, note = choose_api_translation(payload.get("other_names"))
    base["api_status"] = wiki.get("status", "")
    if zh_name:
        confidence = "medium" if zh_name_2 else "high"
        base.update(
            {
                "zh_name": zh_name,
                "zh_name_2": zh_name_2,
                "source": "danbooru_api_other_names",
                "confidence": confidence,
                "note": note,
            }
        )
    else:
        base.update(
            {
                "source": "danbooru_api_no_chinese",
                "confidence": "low",
                "note": wiki.get("error", note),
            }
        )
    return base


def parse_args() -> argparse.Namespace:
    root = Path(__file__).resolve().parent
    output_dir = root / OUTPUT_DIR_NAME
    parser = argparse.ArgumentParser()
    parser.add_argument("--selected", type=Path, default=root / "selected_tags.csv")
    parser.add_argument("--dictionary", type=Path, default=root / "dictionary01.xlsx")
    parser.add_argument("--output-dir", type=Path, default=output_dir)
    parser.add_argument("--base-url", default=os.environ.get("DANBOORU_BASE_URL", "https://danbooru.donmai.us"))
    parser.add_argument("--login", default=os.environ.get("DANBOORU_LOGIN", ""))
    parser.add_argument("--api-key", default=os.environ.get("DANBOORU_API_KEY", ""))
    parser.add_argument("--delay", type=float, default=0.12, help="Seconds between API requests. 0.12 keeps below 10 req/s.")
    parser.add_argument("--timeout", type=int, default=30)
    parser.add_argument("--retry-after", type=float, default=5.0)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--refresh", action="store_true")
    parser.add_argument("--dictionary-only", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.delay < 0.1:
        raise SystemExit("--delay must be at least 0.1 seconds to stay within 10 API requests per second")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    copied_selected = args.output_dir / "selected_tags.csv"
    if args.selected.resolve() != copied_selected.resolve():
        shutil.copy2(args.selected, copied_selected)

    selected_rows = read_csv_dict(args.selected)
    if args.limit:
        selected_rows = selected_rows[: args.limit]
    dictionary = load_dictionary(args.dictionary)
    cache_path = args.output_dir / "danbooru_wiki_cache.json"
    cache = load_cache(cache_path)
    session = requests.Session()

    output_rows: list[dict[str, str]] = []
    api_requests = 0
    for index, row in enumerate(selected_rows, start=1):
        tag = row["name"]
        wiki = None
        if tag not in dictionary and not args.dictionary_only:
            wiki = fetch_wiki(session, tag, args, cache)
            api_requests += 1
            time.sleep(args.delay)
        output_rows.append(build_row(row, dictionary, wiki))
        if index % 100 == 0:
            save_cache(cache_path, cache)
            print(f"processed {index}/{len(selected_rows)}; api_requests={api_requests}")

    save_cache(cache_path, cache)
    fields = [
        "tag_id",
        "name",
        "category",
        "count",
        "zh_name",
        "zh_name_2",
        "source",
        "confidence",
        "note",
        "donmai_url",
        "api_status",
    ]
    output_path = args.output_dir / "selected_tags_full_translation.csv"
    write_csv_dict(output_path, output_rows, fields)

    stats: dict[str, int] = {}
    for row in output_rows:
        stats[row["source"]] = stats.get(row["source"], 0) + 1
    print(
        json.dumps(
            {
                "rows": len(output_rows),
                "api_requests": api_requests,
                "dictionary_entries": len(dictionary),
                "sources": stats,
                "output": str(output_path),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
