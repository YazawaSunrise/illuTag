from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


def norm_tag(value: object) -> str:
    return "" if value is None else str(value).strip()


def load_dictionary_tags(path: Path, sheet_name: str | None = None, tag_column: str = "tag") -> set[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows)
        try:
            tag_index = [norm_tag(cell) for cell in header].index(tag_column)
        except ValueError as exc:
            raise SystemExit(f"Column {tag_column!r} was not found in {path}") from exc

        tags: set[str] = set()
        for row in rows:
            if tag_index < len(row):
                tag = norm_tag(row[tag_index])
                if tag:
                    tags.add(tag)
        return tags
    finally:
        workbook.close()


def write_missing_tags(
    selected_csv: Path, dictionary_xlsx: Path, output_csv: Path, category: str | None
) -> tuple[int, int, int, int]:
    dictionary_tags = load_dictionary_tags(dictionary_xlsx)

    selected_total = 0
    selected_after_category = 0
    missing_total = 0
    output_csv.parent.mkdir(parents=True, exist_ok=True)

    with selected_csv.open("r", encoding="utf-8", newline="") as src, output_csv.open(
        "w", encoding="utf-8-sig", newline=""
    ) as dst:
        reader = csv.DictReader(src)
        writer = csv.writer(dst)
        writer.writerow(["tag", "count"])

        for row in reader:
            selected_total += 1
            if category is not None and norm_tag(row.get("category")) != category:
                continue
            selected_after_category += 1

            tag = norm_tag(row.get("name"))
            if tag and tag not in dictionary_tags:
                writer.writerow([tag, row.get("count", "")])
                missing_total += 1

    return selected_total, selected_after_category, len(dictionary_tags), missing_total


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Find tags present in selected_tags.csv but missing from dictionary01.xlsx."
    )
    parser.add_argument("--selected", type=Path, default=Path("selected_tags.csv"))
    parser.add_argument("--dictionary", type=Path, default=Path("dictionary01.xlsx"))
    parser.add_argument("--output", type=Path, default=Path("selected_tags_missing_from_dictionary01.csv"))
    parser.add_argument("--category", default="0", help="Only include selected_tags rows in this category.")
    args = parser.parse_args()

    category = args.category if args.category != "" else None
    selected_total, selected_after_category, dictionary_total, missing_total = write_missing_tags(
        args.selected, args.dictionary, args.output, category
    )
    print(f"selected_tags rows: {selected_total}")
    print(f"selected_tags rows after category filter: {selected_after_category}")
    print(f"dictionary tags: {dictionary_total}")
    print(f"missing tags written: {missing_total}")
    print(f"output: {args.output}")


if __name__ == "__main__":
    main()
